#!/usr/bin/env python3
import sys 
from openpyxl import load_workbook
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.platypus import Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

def xlsx_a_pdf(input_xlsx, output_pdf):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb.active

    # Leer todas las filas como listas de valores
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    # Configurar PDF landscape
    doc = SimpleDocTemplate(
        output_pdf,
        pagesize=landscape(A4),
        leftMargin=8*mm, rightMargin=8*mm,
        topMargin=8*mm, bottomMargin=8*mm
    )

    styles = getSampleStyleSheet()
    elements = []

    # Identificar secciones: buscar filas que son headers de seccion
    # (ROSES, CARNATIONS, etc.) y filas de header de columnas
    # (FB, FARM, VARIETY, ...)

    # Primero las filas de cabecera (filas 1-7 aprox)
    header_data = []
    data_start = 0
    for i, row in enumerate(all_rows):
        first_cell = str(row[0] or "").strip()
        if first_cell in ("FB", ""):
            # Verificar si es el header de columnas
            if any(str(c or "").strip() == "FARM" for c in row):
                data_start = i
                break
        header_data.append(row)

    # Renderizar cabecera como tabla simple
    if len(header_data) > 0:
        header_table_data = []
        for row in header_data:
            clean_row = [str(c or "") for c in row[:7]]
            header_table_data.append(clean_row)

        if len(header_table_data) > 0:
            ht = Table(header_table_data)
            ht.setStyle(TableStyle([
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTNAME', (3,0), (3,-1), 'Helvetica-Bold'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(ht)
            elements.append(Spacer(1, 4*mm))

    # Procesar secciones de datos
    current_section = []
    section_title = ""
    sections = []

    for i in range(data_start, len(all_rows)):
        row = all_rows[i]
        first_cell = str(row[0] or "").strip()

        # Detectar titulo de seccion (ROSES, CARNATIONS, etc.)
        if first_cell and first_cell not in ("FB", "") and not any(c for c in row[3:15] if c):
            # Es un titulo de seccion
            if current_section:
                sections.append((section_title, current_section))
            section_title = first_cell
            current_section = []
            continue

        # Detectar header de columnas
        if first_cell == "FB" or (any(str(c or "").strip() == "FARM" for c in row)):
            if current_section:
                sections.append((section_title, current_section))
                current_section = []
            current_section.append(row)
            continue

        # Fila de datos o vacia
        if any(c for c in row if c is not None):
            current_section.append(row)

    if current_section:
        sections.append((section_title, current_section))

    # Anchos de columna (en mm) para landscape A4
    # FB(8) FARM(30) VARIETY(30) 12xcm(8 cada) STPR(12) TOTAL(14)
    col_widths = [8*mm, 30*mm, 30*mm] + [8*mm]*12 + [12*mm, 14*mm]

    crimson = colors.Color(0.53, 0.004, 0.055)

    for title, rows in sections:
        if not rows:
            continue

        # Titulo de seccion
        if title:
            p = Paragraph(
                '<b>' + title + '</b>',
                ParagraphStyle('section', fontSize=10,
                    fontName='Helvetica-Bold', spaceAfter=2*mm)
            )
            elements.append(p)

        # Preparar datos de tabla
        table_data = []
        for row in rows:
            clean_row = []
            for j, c in enumerate(row[:17]):
                val = c
                if val is None:
                    val = ""
                elif isinstance(val, float):
                    if j >= 3 and j <= 14:  # columnas cm
                        val = str(int(val)) if val > 0 else ""
                    elif j == 15:  # ST PRICE
                        val = f"{val:.2f}"
                    elif j == 16:  # TOTAL
                        val = f"{val:.2f}"
                    else:
                        if val == int(val):
                            val = str(int(val))
                        else:
                            val = f"{val:.2f}"
                else:
                    val = str(val)
                clean_row.append(val)
            table_data.append(clean_row)

        if not table_data:
            continue

        t = Table(table_data, colWidths=col_widths[:len(table_data[0])])

        style_cmds = [
            ('FONTSIZE', (0,0), (-1,-1), 6.5),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.Color(0.8,0.8,0.8)),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),      # FB center
            ('ALIGN', (3,0), (-1,-1), 'RIGHT'),       # numeros derecha
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]

        # Header row (primera fila) con fondo crimson
        if len(table_data) > 0:
            style_cmds.append(('BACKGROUND', (0,0), (-1,0), crimson))
            style_cmds.append(('TEXTCOLOR', (0,0), (-1,0), colors.white))
            style_cmds.append(('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'))
            style_cmds.append(('FONTSIZE', (0,0), (-1,0), 7))
            style_cmds.append(('ALIGN', (0,0), (-1,0), 'CENTER'))

        # Filas alternas
        for r in range(1, len(table_data)):
            if r % 2 == 0:
                style_cmds.append(
                    ('BACKGROUND', (0,r), (-1,r),
                     colors.Color(0.97,0.97,0.97)))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 6*mm))

    doc.build(elements)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python3 xlsx_a_pdf.py input.xlsx output.pdf")
        sys.exit(1)
    xlsx_a_pdf(sys.argv[1], sys.argv[2])
    print("OK")
